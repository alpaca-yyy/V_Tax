#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
from openpyxl import load_workbook
import configparser
import os


#建立字典对应 编号-企业
#初始化录入的编号为永久编号，用于之后识别不同企业对应的汇总表行数等等

configPath = '.\\config.ini'  #获取配置文件的路径
conf=configparser.ConfigParser()#创建对象用于对配置文件进行操作
conf.read(configPath,encoding="utf-8-sig")#以utf8编码形式读取

time=conf.get("期数设置","time")

if conf.get('初始化','initialized')=='yes':
    print('检测到已初始化，跳过该步骤')
else:
    wb = load_workbook('初始化企业编号.xlsx')
    ws = wb.active

    val = 1
    row1 = 2
    row2 = 2
    dict_players = {}

    for val in range(ws.max_row-1):
        value1 = ws.cell(row=row1,column=1)
        value2 = ws.cell(row=row2,column=2)
        dict_players[value1.value] = value2.value
        val += 1
        row1 += 1
        row2 += 1
    print('初始化完成，所有企业代号：企业名称的字典如下')
    print(dict_players)
    conf.set('初始化','initialized','yes')
    #去除字典中的None值 好但暂时无用
    # def clear_dict(d):
    #     if d is None:
    #         return None
    #     elif isinstance(d, list):
    #         return list(filter(lambda x: x is not None, map(clear_dict, d)))
    #     elif not isinstance(d, dict):
    #         return d
    #     else:
    #         r = dict(
    #                 filter(lambda x: x[1] is not None,
    #                     map(lambda x: (x[0], clear_dict(x[1])),
    #                         d.items())))
    #         if not bool(r):
    #             return None
    #         return r
    # coding=utf-8



    conf.write(open(configPath,'w+',encoding="utf-8-sig")) #将修改写入到配置文件


#建立第X期各企业数据总文件夹
#识别各种类型企业有多少家
if conf.get('初始化',time+'各企业数据')=='True':
    print('检测到'+time+'企业汇总文件夹已生成，跳过此步骤。')
    
else:
    print('现在是'+time)
    os.getcwd()
    path1 = os.getcwd()+'\\'+time+'各企业数据'
    conf.set('初始化','time'+'各企业数据','True')
    os.mkdir(path1)

    #建立各企业二级文件夹
    for key,value in dict_players.items():
        os.mkdir(path1+'\\'+key+value)
    print('累计识别到'+str(len(dict_players))+'家企业，正在生成文件夹……')
    


# In[3]:


#重命名各企业交的报表
import glob
from openpyxl import load_workbook
from xls2xlsx import exchange
#from V_Tax_AutoFill import read_write_sheet


dirs = []
files = []

def traversal_files(path):
    for item in os.scandir(path):
        if item.is_dir():
            dirs.append(item.path)
        elif item.is_file():
            files.append(item.path)


if __name__ == '__main__':
    path = '.\\'+str(time)+'各企业数据'

traversal_files(path)
print('正在整理'+time+'各企业上交的报表')

print('dirs:', dirs)
print('files:', files)


def read_write_sheet(readworksheet1,readworksheet2):
    '''
    读取企业利润表、纳税申报表数据
    思路：1.识别这是哪家企业 2.字典反查企业编码留用
    readworksheet1:利润表的路径
    readworksheet2:纳税申报表的路径
    第几期
    '''
    wb = load_workbook(readworksheet1)
    ws = wb.active
#data_only=True https://blog.csdn.net/claria029/article/details/116486904
    wb2 = load_workbook(readworksheet2,data_only=True)
    ws2 = wb2['汇总表']
    
    NameCell = ws.cell(row=3,column=1)
    Name=str(NameCell.value)[5:30]
#使用 .keys() 和 .values() 在 Python 字典中按值查找键 https://blog.csdn.net/wjj2586590669/article/details/126396851    
    list_of_key = list(dict_players.keys())
    list_of_value = list(dict_players.values())
#警惕 因为使用字典查询，柠檬云生成的报表中编制单位和初始化填入的名称不一样会报错 
    position = list_of_value.index(Name)
    NameCode=list_of_key[position]
    print('您正在读取的是'+'\033[1m'+list_of_key[position]+str(NameCell.value)[5:30]+'\033[0m'+'的数据')
    
    #可以使用配置文件！！！！
    #定位准备输入值在利润表中的行数
    list1 = [5,6,8,9,10,13,15,16,18,20]
    #定位准备输入值在汇总表中的列数
    list2=  [5,6,8,9,10,11,12,13,14,15,16,17,18]
    #定位准备输入值在纳税申报表中的列
    list3=[2,3,4]
    
#之后应该改成自动是识别第几期
    if time == '第一期':
       whichseason = 1
    elif time == '第二期':
       whichseason = 2
    elif time == '第三期':
       whichseason = 3
    elif time == '第四期':
       whichseason = 4
    elif time == '第五期':
       whichseason = 4
    else:
       print("时期不在第一期和第五期之间！")

    #营业收入、营业成本等依序存入列表data_list
    data_list = []

    #读取数据
    for d in list1:
        content = ws.cell(row=d,column=4)
        data_list.append(content.value)

    for f in list3:
        报税表单元格= ws2.cell(row=int(whichseason)+4,column=f)
        data_list.append(报税表单元格.value)
    print(data_list)
    
    '''
    写入data_list的数据到汇总表中
    思路：1.先找出要录入的企业编号对应的行数

    '''
    #写入数据
    wb3 = load_workbook('财务数据汇总表.xlsx')
    ws3 = wb3[time]
#查询企业代码对应的行数 NameCode变量是企业代码 players_row是对应的行数
    for jigou in ws3['C']:
        if jigou.value == NameCode:
            players_row = jigou.row
            print('该企业对应的行数是'+str(jigou.row)+'正在写入中……')
    
    val = 0
    for col in list2:
        ws3.cell(row=players_row,column=col,value=data_list[val])
        print(data_list[val])
        val+=1

    wb3.save('财务数据汇总表.xlsx')
    data_list.clear

for x in dirs:
    print('当前目录是'+os.getcwd())
    print('itemlist_dir的值是'+x)
    企业目录 = os.path.join(os.getcwd(),x[2:])
    print('企业目录是'+企业目录)
    itemlist_dir = x
    
    
#如果企业上交的表文件名中没有“利润”“资产负债”“明细”“纳税”等关键字 会报错
    PathOfPL = glob.glob(os.path.join(itemlist_dir,'*利润*.*'))
#     print(PathOfPL)
    PathOfBS = glob.glob(os.path.join(itemlist_dir,'*资产负债*.*'))
#     print(PathOfBS)
    PathOfSGL = glob.glob(os.path.join(itemlist_dir,'*明细*.*'))
#     print(PathOfSGL)
    PathOfTaxReturnForm = glob.glob(os.path.join(itemlist_dir,'*纳税*.*'))
#     print(PathOfTaxReturnForm)
    
    exchange((
    os.path.abspath
    (time+'各企业数据'+os.path.dirname(str(PathOfPL)[13:]))))
    
    PathOfPL = glob.glob(os.path.join(itemlist_dir,'*利润*.*'))
    PathOfBS = glob.glob(os.path.join(itemlist_dir,'*资产负债*.*'))
    PathOfSGL = glob.glob(os.path.join(itemlist_dir,'*明细*.*'))
    PathOfTaxReturnForm = glob.glob(os.path.join(itemlist_dir,'*纳税*.*'))
    print(PathOfPL)
    print(PathOfBS)
    print(PathOfSGL)
    print(PathOfTaxReturnForm)
    
    if len(PathOfPL) == 0:
        print('该企业的利润表有问题！！')
    else:
        os.rename(PathOfPL[0],itemlist_dir+'\\'+time+'利润表.xlsx')
        PathOfPL2=itemlist_dir+'\\'+time+'利润表.xlsx'
    if len(PathOfBS) == 0:
        print('该企业的资产负债表有问题！！')
    else:
        os.rename(PathOfBS[0],itemlist_dir+'\\'+time+'资产负债表.xlsx')

    if len(PathOfSGL) == 0:
        print('该企业的明细账有问题！！')
    else:
        os.rename(PathOfSGL[0],itemlist_dir+'\\'+time+'明细账.xlsx')

    if len(PathOfTaxReturnForm) == 0:
        print('该企业的纳税申报表有问题！！')
    else:
        os.rename(PathOfTaxReturnForm[0],itemlist_dir+'\\'+time+'纳税申报表.xlsx')

    
    print(glob.glob(os.path.join(str(企业目录), '*.xlsx', '*利润*.*')))

    read_write_sheet(企业目录+'\\'+time+'利润表.xlsx',企业目录+'\\'+time+'纳税申报表.xlsx')


